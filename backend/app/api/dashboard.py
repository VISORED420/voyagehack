"""
Dashboard API Routes
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import date
from io import BytesIO
import logging

from ..models.database import db
from ..models.schemas import DashboardStats
from ..services.cache import cache_service
from ..config import settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/dashboard", tags=["Dashboard"])


@router.get("/event/{event_id}/stats", response_model=DashboardStats)
async def get_dashboard_stats(
    event_id: str,
    bypass_cache: bool = Query(False, description="Bypass cache and fetch fresh data")
):
    """Get real-time dashboard statistics for an event"""
    cache_key = cache_service.key_stats(event_id)

    # Try to get from cache first
    if not bypass_cache:
        cached_stats = await cache_service.get(cache_key)
        if cached_stats:
            logger.debug(f"Cache HIT for stats: {event_id}")
            return cached_stats

    # Cache miss - fetch from database
    event = db.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    stats = db.get_dashboard_stats(event_id)

    # Store in cache
    await cache_service.set(cache_key, stats, settings.CACHE_STATS_TTL)
    logger.debug(f"Cache SET for stats: {event_id}")

    return stats


@router.get("/event/{event_id}")
async def get_full_dashboard(
    event_id: str,
    bypass_cache: bool = Query(False, description="Bypass cache and fetch fresh data")
):
    """Get complete dashboard data including event, stats, and bookings"""
    cache_key = cache_service.key_dashboard(event_id)

    # Try to get from cache first
    if not bypass_cache:
        cached_dashboard = await cache_service.get(cache_key)
        if cached_dashboard:
            logger.debug(f"Cache HIT for dashboard: {event_id}")
            return cached_dashboard

    # Cache miss - fetch from database
    event = db.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    stats = db.get_dashboard_stats(event_id)
    bookings = db.get_bookings(event_id)
    inventory = db.get_inventory(event_id)

    # Format bookings for frontend
    formatted_bookings = []
    for b in bookings:
        formatted_bookings.append({
            "id": b.get("id", ""),
            "ref": b.get("booking_reference", ""),
            "guest": b.get("guest_name", ""),
            "email": b.get("guest_email", ""),
            "phone": b.get("guest_phone", ""),
            "hotel": b.get("hotel", ""),
            "room": b.get("room_type", ""),
            "checkIn": b["check_in_date"].strftime("%d %b") if hasattr(b["check_in_date"], "strftime") else b["check_in_date"],
            "checkOut": b["check_out_date"].strftime("%d %b") if hasattr(b["check_out_date"], "strftime") else b["check_out_date"],
            "nights": b.get("num_nights", 0),
            "rooms": b.get("num_rooms", 1),
            "adults": b.get("num_adults", 2),
            "children": b.get("num_children", 0),
            "total": b.get("total_amount", 0),
            "paid": b.get("amount_paid", 0),
            "due": b.get("amount_due", 0),
            "paymentStatus": b.get("payment_status", "pending"),
            "status": b.get("status", "pending"),
            "category": b.get("category", ""),
            "side": b.get("side", ""),
            "dietary": b.get("dietary_requirements", []),
            "source": b.get("booking_source", "Microsite")
        })

    # Format inventory
    formatted_inventory = []
    for inv in inventory:
        formatted_inventory.append({
            "id": inv["id"],
            "hotel": inv["hotel_name"],
            "room": inv["room_type"],
            "blocked": inv["total_rooms_blocked"],
            "booked": inv["rooms_booked"],
            "available": inv["total_rooms_blocked"] - inv["rooms_booked"],
            "rate": inv["negotiated_rate"],
            "validTill": inv["release_date"].strftime("%b %d") if inv.get("release_date") and hasattr(inv["release_date"], "strftime") else ""
        })

    days_to_event = (event["start_date"] - date.today()).days

    dashboard_data = {
        "event": {
            "id": event["id"],
            "name": event["name"],
            "code": event["event_code"],
            "type": event["event_type"],
            "dates": f"{event['start_date'].strftime('%B %d')}-{event['end_date'].strftime('%d, %Y')}",
            "destination": event["destination"],
            "status": event["status"],
            "daysToEvent": max(0, days_to_event),
            "bookingDeadline": event["booking_deadline"].strftime("%B %d, %Y") if event.get("booking_deadline") else None
        },
        "stats": stats,
        "bookings": formatted_bookings,
        "inventory": formatted_inventory
    }

    # Store in cache
    await cache_service.set(cache_key, dashboard_data, settings.CACHE_DASHBOARD_TTL)
    logger.debug(f"Cache SET for dashboard: {event_id}")

    return dashboard_data


@router.get("/event/{event_id}/export/rooming-list")
async def export_rooming_list(event_id: str, format: str = "excel"):
    """Export rooming list to Excel or CSV"""
    event = db.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    bookings = db.get_bookings(event_id)

    if format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rooming List"

            # Header styling
            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            # Headers
            headers = [
                "Booking Ref", "Guest Name", "Email", "Phone", "Hotel",
                "Room Type", "Check-in", "Check-out", "Nights", "Rooms",
                "Adults", "Children", "Total (₹)", "Paid (₹)", "Due (₹)",
                "Payment Status", "Booking Status", "Category", "Side", "Dietary"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_num, b in enumerate(bookings, 2):
                ws.cell(row=row_num, column=1, value=b.get("booking_reference", ""))
                ws.cell(row=row_num, column=2, value=b.get("guest_name", ""))
                ws.cell(row=row_num, column=3, value=b.get("guest_email", ""))
                ws.cell(row=row_num, column=4, value=b.get("guest_phone", ""))
                ws.cell(row=row_num, column=5, value=b.get("hotel", ""))
                ws.cell(row=row_num, column=6, value=b.get("room_type", ""))
                ws.cell(row=row_num, column=7, value=str(b.get("check_in_date", "")))
                ws.cell(row=row_num, column=8, value=str(b.get("check_out_date", "")))
                ws.cell(row=row_num, column=9, value=b.get("num_nights", 0))
                ws.cell(row=row_num, column=10, value=b.get("num_rooms", 1))
                ws.cell(row=row_num, column=11, value=b.get("num_adults", 2))
                ws.cell(row=row_num, column=12, value=b.get("num_children", 0))
                ws.cell(row=row_num, column=13, value=b.get("total_amount", 0))
                ws.cell(row=row_num, column=14, value=b.get("amount_paid", 0))
                ws.cell(row=row_num, column=15, value=b.get("amount_due", 0))
                ws.cell(row=row_num, column=16, value=b.get("payment_status", ""))
                ws.cell(row=row_num, column=17, value=b.get("status", ""))
                ws.cell(row=row_num, column=18, value=b.get("category", ""))
                ws.cell(row=row_num, column=19, value=b.get("side", ""))
                ws.cell(row=row_num, column=20, value=", ".join(b.get("dietary_requirements", [])))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"{event['event_code']}_rooming_list.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")

    elif format == "csv":
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow([
            "Booking Ref", "Guest Name", "Email", "Phone", "Hotel",
            "Room Type", "Check-in", "Check-out", "Nights", "Rooms",
            "Adults", "Children", "Total", "Paid", "Due",
            "Payment Status", "Booking Status"
        ])

        # Data
        for b in bookings:
            writer.writerow([
                b.get("booking_reference", ""),
                b.get("guest_name", ""),
                b.get("guest_email", ""),
                b.get("guest_phone", ""),
                b.get("hotel", ""),
                b.get("room_type", ""),
                str(b.get("check_in_date", "")),
                str(b.get("check_out_date", "")),
                b.get("num_nights", 0),
                b.get("num_rooms", 1),
                b.get("num_adults", 2),
                b.get("num_children", 0),
                b.get("total_amount", 0),
                b.get("amount_paid", 0),
                b.get("amount_due", 0),
                b.get("payment_status", ""),
                b.get("status", "")
            ])

        output.seek(0)
        filename = f"{event['event_code']}_rooming_list.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'csv'")


@router.get("/event/{event_id}/alerts")
async def get_alerts(
    event_id: str,
    bypass_cache: bool = Query(False, description="Bypass cache and fetch fresh data")
):
    """Get alerts and action items for an event"""
    cache_key = cache_service.key_alerts(event_id)

    # Try to get from cache first
    if not bypass_cache:
        cached_alerts = await cache_service.get(cache_key)
        if cached_alerts:
            logger.debug(f"Cache HIT for alerts: {event_id}")
            return cached_alerts

    event = db.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    bookings = db.get_bookings(event_id)
    inventory = db.get_inventory(event_id)
    guests = db.get_guests(event_id)

    alerts = []

    # Payment pending alerts
    pending_payment_bookings = [b for b in bookings if b["payment_status"] == "pending" and b["status"] == "confirmed"]
    if pending_payment_bookings:
        alerts.append({
            "type": "warning",
            "title": f"{len(pending_payment_bookings)} bookings with payment pending",
            "description": "Payment deadline approaching",
            "action": "Send Reminders",
            "action_type": "send_payment_reminders",
            "count": len(pending_payment_bookings)
        })

    # Pending RSVP alerts
    pending_rsvp = len([g for g in guests if g.get("rsvp_status") == "pending"])
    if pending_rsvp > 0:
        alerts.append({
            "type": "info",
            "title": f"{pending_rsvp} guests haven't responded to invitation",
            "description": "Consider sending a follow-up reminder",
            "action": "Send Follow-up",
            "action_type": "send_rsvp_reminder",
            "count": pending_rsvp
        })

    # Low inventory alerts
    for inv in inventory:
        available = inv["total_rooms_blocked"] - inv["rooms_booked"]
        if available <= 3 and available > 0:
            alerts.append({
                "type": "warning",
                "title": f"Low availability: {inv['room_type']} at {inv['hotel_name']}",
                "description": f"Only {available} rooms remaining",
                "action": "Manage Inventory",
                "action_type": "manage_inventory"
            })

    # Release date alerts
    from datetime import timedelta
    for inv in inventory:
        if inv.get("release_date"):
            days_to_release = (inv["release_date"] - date.today()).days
            if 0 < days_to_release <= 10:
                available = inv["total_rooms_blocked"] - inv["rooms_booked"]
                alerts.append({
                    "type": "success",
                    "title": f"Room block release in {days_to_release} days",
                    "description": f"{available} rooms at {inv['hotel_name']} still available",
                    "action": "Manage Inventory",
                    "action_type": "manage_inventory"
                })

    # Store in cache (alerts change frequently, short TTL)
    await cache_service.set(cache_key, alerts, settings.CACHE_STATS_TTL)
    logger.debug(f"Cache SET for alerts: {event_id}")

    return alerts


# ============================================================
# CACHE MANAGEMENT ENDPOINTS
# ============================================================

@router.post("/event/{event_id}/cache/invalidate")
async def invalidate_event_cache(event_id: str):
    """Invalidate all cached data for a specific event"""
    deleted_count = await cache_service.invalidate_event_cache(event_id)
    return {
        "success": True,
        "message": f"Invalidated cache for event {event_id}",
        "keys_deleted": deleted_count
    }


@router.get("/cache/status")
async def get_cache_status():
    """Get Redis cache status and connection info"""
    return {
        "enabled": settings.REDIS_ENABLED,
        "connected": cache_service.is_connected,
        "host": settings.REDIS_HOST,
        "port": settings.REDIS_PORT,
        "ttl_settings": {
            "default": settings.CACHE_DEFAULT_TTL,
            "dashboard": settings.CACHE_DASHBOARD_TTL,
            "stats": settings.CACHE_STATS_TTL
        }
    }
